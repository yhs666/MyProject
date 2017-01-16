#!/usr/bin/env python
#coding:utf-8
'''
Created on Nov 3, 2016

@author: yang.hongsheng
'''
import sys  
import openpyxl
import copy
from datetime import datetime
from insert_logs import print_logs

DATETIME_FMT = '%Y-%m-%d %H:%M:%S'


inser_issue=[]
updatetime = datetime.strftime(datetime.now(), DATETIME_FMT)
def insert_execl(tickets_info_list):
    xfile = openpyxl.load_workbook('temp.xlsx')

    sheetname ="report"
    sheet = xfile.get_sheet_by_name(sheetname)
    tickets_info = copy.deepcopy(tickets_info_list)
    #execl
    hang =2
    # Insert Execl
    for ticket in tickets_info:
        try:
            #insert update time
            ticket.append(updatetime)
            for k in range(0,len(ticket)):
                if   ticket[k] is None:
                    ticket[k]=""
                elif type(ticket[k]) is datetime:
                    ticket[k]= datetime.strftime(ticket[k],DATETIME_FMT)
                elif type(ticket[k]) is int:
                    pass
                elif type(ticket[k]) is long:
                    pass
                else:
                    ticket[k]=  ticket[k].encode("utf-8") 
                #insert cells
                '''
                try:
                    ticket[k]=int(ticket[k])
                except:
                    pass
                    print "nn"
                '''
                if k==3:
                    ticket[k]=int(ticket[k])
                    
                sheet.cell(row=hang,column= k+1,value= ticket[k])
            
        except Exception,e:
                print "--------------------------Insert Execl have issue.-------------------------------------------------------"
                print_logs(ticket)
                print_logs(str(e))
                print "---------------------------Insert Execl have issue.---------------------------------------------"
                inser_issue.append(ticket[0])
        # insert the next row
        hang=hang+1
    
    rname = "report " + datetime.now().strftime('%Y%m%d %H%M%S') + ".xlsx"
    xfile.save(rname)

    print_logs("Insert execl Done!")

    return inser_issue


def insert_execl_effort(tickets_info_list):
    xfile = openpyxl.load_workbook('effort.xlsx')

    sheetname ="Effort"
    sheet = xfile.get_sheet_by_name(sheetname)
    tickets_info = copy.deepcopy(tickets_info_list)
    #execl
    hang =2
    # Insert Execl
    for ticket in tickets_info:
        try:
            #insert update time
            ticket.append(updatetime)
            for k in range(0,len(ticket)):                    
                sheet.cell(row=hang,column= k+1,value= ticket[k])
            
        except Exception,e:
                print "--------------------------Insert Execl have issue.-------------------------------------------------------"
                print_logs(ticket)
                print_logs(str(e))
                print "---------------------------Insert Execl have issue.---------------------------------------------"
                inser_issue.append(ticket[0])
        # insert the next row
        hang=hang+1
    
    rname = "Effort " + datetime.now().strftime('%Y%m%d %H%M%S') + ".xlsx"
    xfile.save(rname)

    print_logs("Insert execl Done!")

    return inser_issue

            


