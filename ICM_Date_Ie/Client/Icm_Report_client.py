#!/usr/bin/env python
#coding:utf-8
'''
Created on Aug 31, 2016

@author: yang.hongsheng
'''
# -*- coding: utf-8 -*- 

#import Queue

import redis
import logging
import os
import sys
import openpyxl
import json
import hashlib
import time  
import datetime
import autopy


ip ="172.31.4.119"
password = "wasu.com"
time_out = 60

FILE=os.getcwd()
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s:%(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                    datefmt='%a, %d %b %Y %H:%M:%S',
                    filename = os.path.join(FILE,'icm_report_log.txt'),
                    filemode='a')

try:
    myredis = redis.Redis(host=ip,password=password,port = 6379)
    print "Connect Redis OK!"
    logging.info('Connect Redis OK!')

except:
    print "Redis connect issue!"
    logging.info('Redis connect issue!')
    sys.exit()


keyword=[
        'IncidentSeverity',
        'Title',
        'Effort Time',
        'SubType',
        'EscalationOccured',
        'Trigger',
        'Azure Source',
        'Created Date',
        'Mitigated Date',
        'Resolved Date',
        'Ticket State',
        'Sub Status',
        'Owning Service',
        'Owning Team',
        'Impacted Services',
        'Impacted Teams',
        'Impacted Component',
        'Service Responsible',
        'Keywords',
        'Current Summary',
        'updatetime',
         ]
redis_key=[
        'Severity:',
        'Title:',
        'Ops Team Effort:',
        'Azure SubType:',
        'Escalation Status:',
        'Trigger Field:',
        'Azure Source:',
        'Created Date',
        'Time Mitigated (CST):',
        'Time Resolved (CST):',
        'Ticket State',
        'Sub Status:',
        'Owning Service:',
        'Owning Team:',
        'Impacted Services:',
        'Impacted Teams:',
        'Impacted Component:',
        'Service Responsible:',
        'Keywords:',
        'Current Summary:',
        'updatetime'
        ]
# Get icm tickets

icm = open("icm.txt","r")
qs = icm.readline()
q=[]
while qs:
    s = qs.strip()
    if len(s) != 0  and  s[0] != "#"  :
        q.append(s)
    qs = icm.readline()
icm.close()    
print q
q.sort()
logging.info(" ".join(q))


# get submit key
m2 = hashlib.md5()   
m2.update(":".join(q))   
key= m2.hexdigest()
print "----------------------------------"
print "key: ", key
print "----------------------------------"
logging.info("*********************************")
logging.info(key)
logging.info("*********************************")

key1 = "temp_list" 
keystatus = key + ":status"  

#set time line
t =datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
keyt = key + ":time"
myredis.set(keyt,t)

s0 = raw_input("These thickets will submit, Do you want submit? (Y/N) default： Y \n")
if s0 =="N" or s0=="n":
    sys.exit()
else:
    print "submit tickets"
    
if key in myredis.lrange(key1,0,-1):
    pass
    print "Key: ", key, " Had exist!!"
else:
    #set icm tickets to redis
    print "Need insert key:",key
    setredis = myredis.set(key,json.dumps(q))
    if not setredis:
        msg = "Insert redis  icm tickets have issue"
        print msg
        logging.info(msg)
 
        
if myredis.get(keystatus) =="done":
    s = raw_input("These thickets data had in redis, Do you want run again? (Y/N) default： N \n")
    if s =="Y" or s=="y":
        myredis.rpush(key1,key)
        myredis.set(keystatus,"submit")
elif myredis.get(keystatus) =="running":
    pass
    print "Don't need submit these tickets, is running!"
else:
    myredis.rpush(key1,key)
    myredis.set(keystatus,"submit")        
#myredis.set(keystatus,"done")
print "Waiting get the icm date:",

t_time = 1
while 1:
    time.sleep(1)
    k =myredis.get(keystatus)
    if   k =="done":
        break
    elif k == "running":
        t_time=t_time+1
        print "!",
    else:
        print ".",

 
print ""
print "Used second time: ", t_time
   
inser_issue=[]
date_issue=[]
#execl
xfile = openpyxl.load_workbook('temp.xlsx')

sheet = xfile.get_sheet_by_name("report")
hang =2

for i in json.loads(myredis.get(key)):
    key = i +":detail:status"
    
    if myredis.get(key) == "ok" :
        key = i +":detail"
        icm_detail= json.loads(myredis.get(key))
        insert_db=[]
        insert_db.append(i)
        for j in range(0,len(keyword)):
            insert_db.append(icm_detail[redis_key[j]].encode('utf-8'))
            
        #print i,insert_db   
        #msg = i + "  " + " ".join(insert_db)   
        logging.info(i )
        

        try:
            #insert execl
            insert_db.insert(7, "")
            for k in range(0,len(insert_db)):

                if k < 2 or k==3 :
                    sheet.cell(row=hang,column= k+1,value= int(insert_db[k])) 
                    _cell = sheet.cell(row=hang,column= k+1)
                    _cell.number_format = '0'
                         
                elif k != 7  :
                    sheet.cell(row=hang,column= k+1,value= insert_db[k])
                else:
                    try:
                        if insert_db[k+1] =="CSS-ASMS":
                            v = "WASMS"
                        elif  insert_db[k+1] in ["CSS-IAAS Platform","CSS-Networking","CSS-IAAS Availability"]:
                            v = "WATS"
                        elif  insert_db[k+1] in ["CSS-SQL","CSS-Developer","CSS-WebApp"]: 
                            v = "CIE" 
                        elif  insert_db[k+1] in ["CSS-Shanghai"]: 
                            v = "CSS" 
                        elif  insert_db[k+1] in ["Component Team","First Party"]: 
                            v = "Component"
                        else:
                            v ="" 
                    except:
                        v=""
                    sheet.cell(row=hang,column= k+1,value= v)
                    #sheet.cell(row=hang,column= k+2,value= insert_db[k])
  
            hang=hang+1
        except Exception,e:
            msg = "add  %s : %s" %(k,hang) 
            print "---------------------------------------------------------------------------------"
            print msg
            logging.info(msg)
            inser_issue.append(i)
    else:
        print "---------------------------------------------------------------------------------"
        print i, "Get Date have issue! Please get this tickets again!!"
        date_issue.append(i)




print "---------------------------------------------------------------------------------"
errorlog = "ICM tickets Data have issue:  "  + "   ".join(date_issue)
print errorlog
logging.info(errorlog)

insertlog = "ICM tickets Insert Data have issue:  "  + "   ".join(inser_issue)
print insertlog
logging.info(insertlog)
rname = "report " + datetime.datetime.now().strftime('%Y-%m-%d %H %M %S') + ".xlsx"
xfile.save(rname)
print "---------------------------------------------------------------------------------"
print "Done!!!!!!!!!"
logging.info("Run done!")
#autopy.alert.alert(msg="Had Done the work!",title="ICM Report",)
autopy.alert.alert(msg="ICM Report",title="Had Done the work")
sys.exit()



